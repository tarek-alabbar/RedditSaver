import praw
import os
import requests
import pandas as pd
import pathlib
from urllib.parse import urlparse
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font

# Replace the following with your credentials
client_id = "RycAI5jwR7OsVjhfUyLqvg"
client_secret = "l_WcCKlWDhHvaubmaXbhHi4IyxQ2KQ"
user_agent = "redditdevScraper"
folder_path = pathlib.Path().resolve()

reddit = praw.Reddit(client_id=client_id, client_secret=client_secret, user_agent=user_agent)

def download_file(url, folder, filename):
    response = requests.get(url)
    os.makedirs(folder, exist_ok=True)
    with open(os.path.join(folder, filename), 'wb') as f:
        f.write(response.content)

def save_user_posts(user):
    reddit_user = reddit.redditor(user)
    user_folder = user
    counter = 1
    
    log_data = []

    for submission in reddit_user.submissions.new(limit=None):
        post_id = submission.id
        title = submission.title.replace('/', '_')
        print(f'{counter}. Downloading {post_id}...')
        counter += 1
        
        if submission.is_video:
            post_type = 'clip'
            video_url = submission.media['reddit_video']['fallback_url']
            local_file = os.path.join(user_folder, 'clips', f'{post_id}.mp4')
            download_file(video_url, os.path.join(user_folder, 'clips'), f'{post_id}.mp4')
        elif submission.url.endswith(('.jpg', '.jpeg', '.png', '.gif')):
            post_type = 'image'
            img_ext = os.path.splitext(urlparse(submission.url).path)[1]
            local_file = os.path.join(user_folder, 'images', f'{post_id}{img_ext}')
            download_file(submission.url, os.path.join(user_folder, 'images'), f'{post_id}{img_ext}')
        else:
            post_type = 'text'
            local_file = os.path.join(user_folder, 'texts', f'{post_id}.txt')
            os.makedirs(os.path.join(user_folder, 'texts'), exist_ok=True)
            with open(os.path.join(user_folder, 'texts', f'{post_id}.txt'), 'w', encoding='utf-8') as f:
                f.write(submission.selftext)

        log_data.append({
            'post_id': post_id,
            'post_title': title,
            'post_type': post_type,
            'link': f'=HYPERLINK("{os.path.join(folder_path, local_file)}")'
        })

    df = pd.DataFrame(log_data)
    os.makedirs(user_folder, exist_ok=True)
    file_path = os.path.join(user_folder, 'UserLog.xlsx')
    df.to_excel(file_path, index=False)

    # Make the links in the 'link' column clickable
    wb = load_workbook(file_path)
    ws = wb.active
    for r in ws.iter_rows(min_row=2, max_col=4, max_row=len(df)+1):
        r[3].style = 'Hyperlink'

    # Save the modified Excel file
    wb.save(file_path)

if __name__ == '__main__':
    user = input('Enter the Reddit username: ')
    save_user_posts(user)